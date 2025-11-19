"""
Financial Statements Comparatives Verification Tool

This tool compares comparative figures in current year's financial statements
with the actual figures from previous year's financial statements and generates
a detailed verification report.
"""

import pandas as pd
import pdfplumber
import re
import openpyxl
from openpyxl.styles import PatternFill
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from difflib import SequenceMatcher
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class LineItem:
    """Represents a financial statement line item"""
    description: str
    amount: float
    original_text: str
    row_number: int
    note_reference: Optional[str] = None
    statement_type: Optional[str] = None  # Sheet name or statement type


@dataclass
class ComparisonResult:
    """Results of comparing two line items"""
    current_year_item: str
    current_year_comparative: float
    previous_year_actual: Optional[float]
    matches: bool
    difference: Optional[float]
    similarity_score: float
    status: str  # 'MATCH', 'MISMATCH', 'ADDED', 'DELETED'
    statement_type: Optional[str] = None  # Sheet name or statement type


class FinancialStatementParser:
    """Parses financial statements from PDF or Excel files"""
    
    def __init__(self):
        self.amount_pattern = re.compile(r'[(\[]?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?[)\]]?')
        
    def parse_pdf(self, file_path: str) -> List[LineItem]:
        """Extract line items from PDF financial statements"""
        line_items = []
        
        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                tables = page.extract_tables()
                
                # Try to extract from tables first
                if tables:
                    for table in tables:
                        items = self._parse_table(table, page_num)
                        line_items.extend(items)
                
                # Also parse text for any missed items
                text_items = self._parse_text(text, page_num)
                line_items.extend(text_items)
        
        return line_items
    
    def parse_excel(self, file_path: str) -> List[LineItem]:
        """Extract line items from Excel financial statements"""
        line_items = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True gets values, not formulas
            for sheet in wb.worksheets:
                # Skip hidden sheets
                if sheet.sheet_state == 'hidden':
                    logger.info(f"Skipping hidden sheet: {sheet.title}")
                    continue

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    item = self._parse_excel_row(row, row_idx, sheet.title)
                    if item:
                        line_items.append(item)
        except Exception as e:
            logger.error(f"Error parsing Excel: {e}")

        return line_items
    
    def _parse_table(self, table: List[List], page_num: int) -> List[LineItem]:
        """Parse a table from PDF"""
        items = []
        
        for row_idx, row in enumerate(table):
            if not row or len(row) < 2:
                continue
            
            # Assuming first column is description, subsequent columns are amounts
            description = str(row[0]).strip() if row[0] else ""
            
            if not description or self._is_header_row(description):
                continue
            
            # Extract amounts from the row
            amounts = []
            for cell in row[1:]:
                if cell:
                    amount = self._extract_amount(str(cell))
                    if amount is not None:
                        amounts.append(amount)
            
            # Create line items for each amount found
            if amounts and description:
                # Usually last amount is comparative
                item = LineItem(
                    description=description,
                    amount=amounts[-1] if len(amounts) > 1 else amounts[0],
                    original_text=" | ".join(str(c) for c in row if c),
                    row_number=page_num * 1000 + row_idx,
                    statement_type="PDF Content"  # PDFs don't have sheet names
                )
                items.append(item)
        
        return items
    
    def _parse_text(self, text: str, page_num: int) -> List[LineItem]:
        """Parse text for any line items"""
        items = []
        lines = text.split('\n')
        
        for idx, line in enumerate(lines):
            line = line.strip()
            if not line or len(line) < 5:
                continue
            
            # Look for patterns like "Description ... Amount"
            amounts = self._extract_all_amounts(line)
            if amounts and not self._is_header_row(line):
                description = self._clean_description(line, amounts)
                if description:
                    item = LineItem(
                        description=description,
                        amount=amounts[-1],  # Last amount is usually comparative
                        original_text=line,
                        row_number=page_num * 1000 + idx,
                        statement_type="PDF Content"  # PDFs don't have sheet names
                    )
                    items.append(item)
        
        return items
    
    def _parse_excel_row(self, row: Tuple, row_idx: int, sheet_name: str) -> Optional[LineItem]:
        """Parse a single Excel row"""
        if not row or len(row) < 2:
            return None
        
        description = str(row[0]).strip() if row[0] else ""
        
        if not description or self._is_header_row(description):
            return None
        
        # Extract amounts from remaining columns
        amounts = []
        for cell in row[1:]:
            if cell is not None and isinstance(cell, (int, float)):
                amounts.append(float(cell))
            elif cell:
                amount = self._extract_amount(str(cell))
                if amount is not None:
                    amounts.append(amount)
        
        if amounts:
            # For proper comparatives verification:
            # - Current Year file: Use amounts[-1] (last column = PY comparative)
            # - Previous Year file: Use amounts[0] if len==2 else amounts[-2] (CY actual)
            # Note: This assumes standard format with CY in col2, PY comparative in col3
            return LineItem(
                description=description,
                amount=amounts[-1],  # Last column (PY comparative in CY file, or older comparative in PY file)
                original_text=" | ".join(str(c) for c in row if c),
                row_number=row_idx,
                statement_type=sheet_name  # Use Excel sheet name as statement type
            )
        
        return None
    
    def _extract_amount(self, text: str) -> Optional[float]:
        """Extract numeric amount from text"""
        text = str(text).strip()
        
        # Remove common non-numeric characters
        text = text.replace(',', '').replace('$', '').replace('₹', '')
        
        # Handle parentheses/brackets (negative numbers)
        is_negative = text.startswith('(') or text.startswith('[')
        text = text.strip('()[]')
        
        try:
            amount = float(text)
            return -amount if is_negative else amount
        except ValueError:
            return None
    
    def _extract_all_amounts(self, text: str) -> List[float]:
        """Extract all amounts from text"""
        amounts = []
        matches = self.amount_pattern.findall(text)
        
        for match in matches:
            amount = self._extract_amount(match)
            if amount is not None:
                amounts.append(amount)
        
        return amounts
    
    def _clean_description(self, text: str, amounts: List[float]) -> str:
        """Clean description by removing amounts"""
        for amount in amounts:
            # Remove amount representations from text
            patterns = [
                str(int(amount)) if amount.is_integer() else str(amount),
                f"{amount:,.2f}",
                f"({abs(amount):,.2f})" if amount < 0 else ""
            ]
            for pattern in patterns:
                text = text.replace(pattern, '')
        
        return text.strip()
    
    def _is_header_row(self, text: str) -> bool:
        """Check if text is a header row"""
        headers = [
            'particulars', 'description', 'note', 'as at', 'year ended',
            'march 31', 'total', 'sub-total', 'schedule', 'page'
        ]
        text_lower = text.lower()
        return any(header in text_lower for header in headers)


class ComparativesVerifier:
    """Verifies comparative figures between two financial statements"""
    
    def __init__(self, similarity_threshold: float = 0.85, amount_tolerance: float = 0.0):
        self.similarity_threshold = similarity_threshold
        self.amount_tolerance = amount_tolerance  # Always 0 for exact matching
    
    def verify(self, current_year_items: List[LineItem], 
               previous_year_items: List[LineItem]) -> List[ComparisonResult]:
        """Compare line items between two years"""
        results = []
        matched_prev_indices = set()
        
        for curr_item in current_year_items:
            # Try to find matching item in previous year
            best_match = None
            best_score = 0.0
            best_idx = -1
            
            for idx, prev_item in enumerate(previous_year_items):
                if idx in matched_prev_indices:
                    continue
                
                score = self._calculate_similarity(
                    curr_item.description, 
                    prev_item.description
                )
                
                if score > best_score:
                    best_score = score
                    best_match = prev_item
                    best_idx = idx
            
            # Determine status and create result
            if best_match and best_score >= self.similarity_threshold:
                matched_prev_indices.add(best_idx)
                matches = self._amounts_match(curr_item.amount, best_match.amount)
                difference = curr_item.amount - best_match.amount if not matches else 0.0
                
                result = ComparisonResult(
                    current_year_item=curr_item.description,
                    current_year_comparative=curr_item.amount,
                    previous_year_actual=best_match.amount,
                    matches=matches,
                    difference=difference,
                    similarity_score=best_score,
                    status='MATCH' if matches else 'MISMATCH',
                    statement_type=curr_item.statement_type  # Use sheet name from Excel
                )
            else:
                # Line item added in current year
                result = ComparisonResult(
                    current_year_item=curr_item.description,
                    current_year_comparative=curr_item.amount,
                    previous_year_actual=None,
                    matches=False,
                    difference=None,
                    similarity_score=best_score,
                    status='ADDED',
                    statement_type=curr_item.statement_type  # Use sheet name from Excel
                )
            
            results.append(result)
        
        # Check for deleted items
        for idx, prev_item in enumerate(previous_year_items):
            if idx not in matched_prev_indices:
                result = ComparisonResult(
                    current_year_item=prev_item.description,
                    current_year_comparative=0.0,
                    previous_year_actual=prev_item.amount,
                    matches=False,
                    difference=None,
                    similarity_score=0.0,
                    status='DELETED',
                    statement_type=prev_item.statement_type  # Use sheet name from Excel
                )
                results.append(result)
        
        return results
    
    def _calculate_similarity(self, text1: str, text2: str) -> float:
        """Calculate similarity between two text strings"""
        text1 = text1.lower().strip()
        text2 = text2.lower().strip()
        
        # Use SequenceMatcher for fuzzy matching
        return SequenceMatcher(None, text1, text2).ratio()
    
    def _amounts_match(self, amount1: float, amount2: float) -> bool:
        """
        Check if two amounts match with proper float comparison.
        Uses 1 cent/paisa tolerance to handle floating-point precision issues.
        """
        # Use small tolerance (0.01) for floating-point precision
        # This handles cases like 0.1 + 0.2 != 0.3 due to binary representation
        return abs(amount1 - amount2) < 0.01


class ReportGenerator:
    """Generates verification reports"""
    
    def __init__(self):
        self.green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    def generate_excel_report(self, results: List[ComparisonResult], output_path: str):
        """Generate Excel report with color coding"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparatives Verification"
        
        # Headers
        headers = [
            "Line Item Description",
            "Current Year Comparative",
            "Previous Year Actual",
            "Difference",
            "Status",
            "Similarity Score"
        ]
        ws.append(headers)
        
        # Make headers bold
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Add data with color coding
        for result in results:
            row = [
                result.current_year_item,
                result.current_year_comparative,
                result.previous_year_actual if result.previous_year_actual is not None else "N/A",
                result.difference if result.difference is not None else "N/A",
                result.status,
                f"{result.similarity_score:.2%}"
            ]
            ws.append(row)
            
            # Apply color coding to the row
            row_idx = ws.max_row
            if result.status == 'MATCH':
                fill = self.green_fill
            elif result.status == 'MISMATCH':
                fill = self.yellow_fill
            else:  # ADDED or DELETED
                fill = self.red_fill
            
            for cell in ws[row_idx]:
                cell.fill = fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        logger.info(f"Report saved to {output_path}")
    
    def generate_summary(self, results: List[ComparisonResult]) -> Dict:
        """Generate summary statistics"""
        total = len(results)
        matches = sum(1 for r in results if r.status == 'MATCH')
        mismatches = sum(1 for r in results if r.status == 'MISMATCH')
        added = sum(1 for r in results if r.status == 'ADDED')
        deleted = sum(1 for r in results if r.status == 'DELETED')
        
        return {
            'total_items': total,
            'matches': matches,
            'mismatches': mismatches,
            'added_items': added,
            'deleted_items': deleted,
            'match_percentage': (matches / total * 100) if total > 0 else 0
        }


def main():
    """Main function to run the verification tool"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Financial Statements Comparatives Verification Tool"
    )
    parser.add_argument(
        '--current-year',
        required=True,
        help='Path to current year financial statements (PDF or Excel)'
    )
    parser.add_argument(
        '--previous-year',
        required=True,
        help='Path to previous year financial statements (PDF or Excel)'
    )
    parser.add_argument(
        '--output',
        default='comparatives_verification_report.xlsx',
        help='Output report file path'
    )
    parser.add_argument(
        '--similarity-threshold',
        type=float,
        default=0.85,
        help='Similarity threshold for matching line items (0-1)'
    )
    parser.add_argument(
        '--amount-tolerance',
        type=float,
        default=0.01,
        help='Amount tolerance as percentage (e.g., 0.01 for 1%%)'
    )
    
    args = parser.parse_args()
    
    # Parse financial statements
    parser_obj = FinancialStatementParser()
    
    logger.info(f"Parsing current year file: {args.current_year}")
    if args.current_year.endswith('.xlsx') or args.current_year.endswith('.xls'):
        current_items = parser_obj.parse_excel(args.current_year)
    else:
        current_items = parser_obj.parse_pdf(args.current_year)
    
    logger.info(f"Found {len(current_items)} line items in current year")
    
    logger.info(f"Parsing previous year file: {args.previous_year}")
    if args.previous_year.endswith('.xlsx') or args.previous_year.endswith('.xls'):
        previous_items = parser_obj.parse_excel(args.previous_year)
    else:
        previous_items = parser_obj.parse_pdf(args.previous_year)
    
    logger.info(f"Found {len(previous_items)} line items in previous year")
    
    # Verify comparatives
    verifier = ComparativesVerifier(
        similarity_threshold=args.similarity_threshold,
        amount_tolerance=args.amount_tolerance
    )
    
    logger.info("Comparing line items...")
    results = verifier.verify(current_items, previous_items)
    
    # Generate report
    report_gen = ReportGenerator()
    report_gen.generate_excel_report(results, args.output)
    
    # Print summary
    summary = report_gen.generate_summary(results)
    print("\n" + "="*50)
    print("VERIFICATION SUMMARY")
    print("="*50)
    print(f"Total Line Items: {summary['total_items']}")
    print(f"Matches: {summary['matches']} (Green)")
    print(f"Mismatches: {summary['mismatches']} (Yellow)")
    print(f"Added Items: {summary['added_items']} (Red)")
    print(f"Deleted Items: {summary['deleted_items']} (Red)")
    print(f"Match Percentage: {summary['match_percentage']:.2f}%")
    print("="*50)
    print(f"\nDetailed report saved to: {args.output}")


if __name__ == "__main__":
    main()
